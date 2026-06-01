"""
Examination Seating Arrangement System  v2
------------------------------------------
Features:
  • Subject-wise seating allocation
  • Invigilator assignment per hall
  • Color-formatted Excel report

Requirements:
    pip install pandas openpyxl
"""

import pandas as pd
import random
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────
# 1.  SAMPLE DATA  (replace with your actual data)
# ──────────────────────────────────────────────

random.seed(42)

SUBJECTS = ["Mathematics", "Physics", "Chemistry", "English", "Computer Science"]

students = [
    {
        "roll_no":  f"S{str(i).zfill(3)}",
        "name":     f"Student {i}",
        "branch":   random.choice(["CSE", "ECE", "MECH"]),
        "subject":  random.choice(SUBJECTS),
    }
    for i in range(1, 41)
]

halls = [
    {"hall_id": "H1", "hall_name": "Hall A", "capacity": 15},
    {"hall_id": "H2", "hall_name": "Hall B", "capacity": 15},
    {"hall_id": "H3", "hall_name": "Hall C", "capacity": 15},
]

invigilators = [
    {"name": "Dr. Sharma",  "hall_id": "H1"},
    {"name": "Prof. Reddy", "hall_id": "H2"},
    {"name": "Ms. Priya",   "hall_id": "H3"},
]

# ──────────────────────────────────────────────
# 2.  SEATING ALLOCATION  (subject-wise sorted)
# ──────────────────────────────────────────────

def allocate_seats(students, halls):
    total_capacity = sum(h["capacity"] for h in halls)
    if len(students) > total_capacity:
        raise ValueError(f"Not enough seats! Students: {len(students)}, Capacity: {total_capacity}")

    # Sort by subject so same-subject students sit together per hall
    sorted_students = sorted(students, key=lambda s: s["subject"])

    invig_map = {inv["hall_id"]: inv["name"] for inv in invigilators}

    records, idx = [], 0
    for hall in halls:
        for seat_no in range(1, hall["capacity"] + 1):
            if idx >= len(sorted_students):
                break
            s = sorted_students[idx]
            records.append({
                "Roll No":      s["roll_no"],
                "Name":         s["name"],
                "Branch":       s["branch"],
                "Subject":      s["subject"],
                "Hall":         hall["hall_name"],
                "Hall ID":      hall["hall_id"],
                "Seat No":      seat_no,
                "Invigilator":  invig_map.get(hall["hall_id"], "TBD"),
            })
            idx += 1

    return records

# ──────────────────────────────────────────────
# 3.  EXCEL STYLING HELPERS
# ──────────────────────────────────────────────

# Color palette
COLORS = {
    "header_bg":    "1F4E79",   # dark blue
    "header_font":  "FFFFFF",   # white
    "title_bg":     "2E75B6",   # medium blue
    "alt_row":      "DEEAF1",   # light blue
    "subject_bg": {             # per-subject highlight
        "Mathematics":      "FFF2CC",
        "Physics":          "E2EFDA",
        "Chemistry":        "FCE4D6",
        "English":          "EDEDED",
        "Computer Science": "E8D5F5",
    },
    "summary_header": "375623",
    "invig_bg":     "FFF2CC",
}

def thin_border():
    s = Side(style="thin", color="B0B0B0")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_header_style(cell, bg=COLORS["header_bg"], font_color=COLORS["header_font"]):
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(bold=True, color=font_color, name="Arial", size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = thin_border()

def apply_data_style(cell, bg="FFFFFF", bold=False):
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(bold=bold, name="Arial", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = thin_border()

def auto_col_width(ws, min_w=10, max_w=35):
    for col in ws.columns:
        width = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(width + 4, min_w), max_w)

# ──────────────────────────────────────────────
# 4.  EXPORT TO EXCEL  (colored, multi-sheet)
# ──────────────────────────────────────────────

def export_to_excel(df: pd.DataFrame, filename: str = "seating_arrangement_v2.xlsx"):
    exam_date = date.today().strftime("%d %B %Y")

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        # ── Sheet 1: All Students ──
        df.drop(columns=["Hall ID"]).to_excel(writer, sheet_name="All Students", index=False, startrow=2)

        # ── Sheet per Hall ──
        for hall_name, group in df.groupby("Hall"):
            hall_id   = group["Hall ID"].iloc[0]
            invig     = group["Invigilator"].iloc[0]
            cap       = next(h["capacity"] for h in halls if h["hall_id"] == hall_id)
            seated    = len(group)
            sheet_df  = group.drop(columns=["Hall ID"]).reset_index(drop=True)
            sheet_df.to_excel(writer, sheet_name=hall_name, index=False, startrow=4)

            ws = writer.sheets[hall_name]

            # Title block
            ws.merge_cells("A1:H1")
            ws["A1"] = f"EXAMINATION SEATING ARRANGEMENT — {hall_name}  |  {exam_date}"
            ws["A1"].fill      = PatternFill("solid", fgColor=COLORS["title_bg"])
            ws["A1"].font      = Font(bold=True, color="FFFFFF", size=13, name="Arial")
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 28

            ws.merge_cells("A2:D2")
            ws["A2"] = f"Invigilator: {invig}"
            ws["A2"].fill      = PatternFill("solid", fgColor=COLORS["invig_bg"])
            ws["A2"].font      = Font(bold=True, size=11, name="Arial")
            ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

            ws.merge_cells("E2:H2")
            ws["E2"] = f"Capacity: {cap}   |   Seated: {seated}   |   Vacant: {cap - seated}"
            ws["E2"].font      = Font(bold=True, size=11, name="Arial")
            ws["E2"].alignment = Alignment(horizontal="right", vertical="center")
            ws.row_dimensions[2].height = 22

            # Header row (row 5 = startrow 4 + 1)
            for cell in ws[5]:
                apply_header_style(cell)
            ws.row_dimensions[5].height = 20

            # Data rows with subject colour + alternating rows
            for row_idx, row in enumerate(ws.iter_rows(min_row=6, max_row=5 + seated), start=0):
                subject_val = row[3].value          # "Subject" column (D)
                subj_color  = COLORS["subject_bg"].get(subject_val, "FFFFFF")
                alt_color   = COLORS["alt_row"] if row_idx % 2 == 0 else "FFFFFF"
                for cell in row:
                    # Subject column gets subject color; others alternate
                    bg = subj_color if cell.column == 4 else alt_color
                    apply_data_style(cell, bg=bg)

            auto_col_width(ws)

        # ── Sheet: Hall Summary ──
        summary_data = []
        for h in halls:
            invig = next((inv["name"] for inv in invigilators if inv["hall_id"] == h["hall_id"]), "TBD")
            seated = len(df[df["Hall ID"] == h["hall_id"]])
            subjects_in_hall = ", ".join(sorted(df[df["Hall ID"] == h["hall_id"]]["Subject"].unique()))
            summary_data.append({
                "Hall":            h["hall_name"],
                "Invigilator":     invig,
                "Capacity":        h["capacity"],
                "Students Seated": seated,
                "Vacant Seats":    h["capacity"] - seated,
                "Subjects":        subjects_in_hall,
            })
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name="Hall Summary", index=False, startrow=2)

        ws = writer.sheets["Hall Summary"]
        ws.merge_cells("A1:F1")
        ws["A1"] = f"HALL SUMMARY  |  {exam_date}"
        ws["A1"].fill      = PatternFill("solid", fgColor=COLORS["summary_header"])
        ws["A1"].font      = Font(bold=True, color="FFFFFF", size=13, name="Arial")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        for cell in ws[3]:
            apply_header_style(cell, bg=COLORS["summary_header"])

        for row_idx, row in enumerate(ws.iter_rows(min_row=4, max_row=3 + len(summary_data)), start=0):
            bg = COLORS["alt_row"] if row_idx % 2 == 0 else "FFFFFF"
            for cell in row:
                apply_data_style(cell, bg=bg)

        auto_col_width(ws)

        # ── Style the "All Students" sheet ──
        ws = writer.sheets["All Students"]
        ws.merge_cells("A1:H1")
        ws["A1"] = f"COMPLETE STUDENT SEATING LIST  |  {exam_date}"
        ws["A1"].fill      = PatternFill("solid", fgColor=COLORS["title_bg"])
        ws["A1"].font      = Font(bold=True, color="FFFFFF", size=13, name="Arial")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        for cell in ws[3]:
            apply_header_style(cell)

        for row_idx, row in enumerate(ws.iter_rows(min_row=4, max_row=3 + len(df)), start=0):
            subject_val = row[3].value
            subj_color  = COLORS["subject_bg"].get(subject_val, "FFFFFF")
            alt_color   = COLORS["alt_row"] if row_idx % 2 == 0 else "FFFFFF"
            for cell in row:
                bg = subj_color if cell.column == 4 else alt_color
                apply_data_style(cell, bg=bg)

        auto_col_width(ws)

    print(f"✅  Excel report saved → {filename}")

# ──────────────────────────────────────────────
# 5.  EXPORT TO CSV
# ──────────────────────────────────────────────

def export_to_csv(df: pd.DataFrame, filename: str = "seating_arrangement_v2.csv"):
    df.to_csv(filename, index=False)
    print(f"✅  CSV report saved  → {filename}")

# ──────────────────────────────────────────────
# 6.  CONSOLE PRINT
# ──────────────────────────────────────────────

def print_seating(df: pd.DataFrame):
    print(f"\n{'='*65}")
    print(f"  EXAMINATION SEATING ARRANGEMENT  |  {date.today()}")
    print(f"{'='*65}")
    for hall, group in df.groupby("Hall"):
        invig = group["Invigilator"].iloc[0]
        print(f"\n📋  {hall}  ({len(group)} students)  |  Invigilator: {invig}")
        print("-" * 65)
        print(f"{'Seat':<6} {'Roll No':<10} {'Name':<15} {'Branch':<8} {'Subject'}")
        print("-" * 65)
        for _, row in group.iterrows():
            print(f"{row['Seat No']:<6} {row['Roll No']:<10} {row['Name']:<15} {row['Branch']:<8} {row['Subject']}")
    print(f"\nTotal students seated: {len(df)}")
    print(f"{'='*65}\n")

# ──────────────────────────────────────────────
# 7.  MAIN
# ──────────────────────────────────────────────

if __name__ == "__main__":
    records = allocate_seats(students, halls)
    df      = pd.DataFrame(records)

    print_seating(df)
    export_to_excel(df)
    export_to_csv(df)
